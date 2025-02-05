# DTCDatabase/connection.py

from motor.motor_asyncio import AsyncIOMotorClient
import os
import openpyxl
from .schema import create_schema_validation, create_indexes
from api.config import DatabaseConfig

class DTCDatabase:
    client: AsyncIOMotorClient = None
    db = None
    collection = None
    db_name = DatabaseConfig.name


    @classmethod
    def count_excel_rows(cls):
        """Count the number of rows in the Excel file (excluding header)"""
        try:
            excel_path = os.path.join('api', 'database', 'dtc_descriptions', 'dtc_descriptions.xlsx')
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            # print(workbook)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            print(headers)
            # Subtract 1 to exclude header row
            return sheet.max_row - 1
        except Exception as e:
            print(f"Error counting Excel rows: {str(e)}")
            raise

    @classmethod
    async def import_excel_data(cls, force_update=False):
        """Import DTC codes from Excel file"""
        try:            
            excel_path = os.path.join('api', 'database', 'dtc_descriptions', 'dtc_descriptions.xlsx')
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            # Get first sheet
            sheet = workbook.active
            data = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2):
                row_data = {}
                for header, cell in zip(headers, row):
                    row_data[header] = str(cell.value) if cell.value is not None else ""
                data.append(row_data)
            
            documents = []
            skipped_docs = []
            # print(row)
            for row in data:
                try:
                    document = {
                        "Name": row.get("Name", ""),
                        "Title": row.get("Title", ""),
                        "DTC": row.get("DTC", ""),
                        "Component": row.get("Component", ""),
                        "SEVERITY": row.get("SEVERITY\n(Critical Y/N)", "").lower(),  # Convert to lowercase
                        "Driver_reaction": row.get("Driver reaction", ""),
                        "Test_Condition": row.get("Test Condition", ""),
                        "Fault_Detection": row.get("Fault Detection", ""),
                        "Performance_Limiter": row.get("Performance Limiter", ""),
                        "Residual_torque": row.get("Residual torque [%]", ""),
                        "RED_LAMP": row.get("RED LAMP", ""),
                        "Amber_Lamp": row.get("Amber Lamp", ""),
                        "MIL": row.get("MIL", ""),
                        "Validation": row.get("Validation (MIL ON)", ""),
                        "Healing": row.get("Healing (MIL OFF)", "")
                    }
                    documents.append(document)
                except Exception as e:
                    skipped_docs.append({"DTC": row.get("DTC", ""), "reason": str(e)})
            
            if documents:
                if force_update:
                    # Clear existing data
                    await cls.collection.delete_many({})
                    print("Cleared existing DTC data from database")
                
                # Insert all documents at once
                try:
                    result = await cls.collection.insert_many(documents, ordered=False)
                    successful_inserts = len(result.inserted_ids)
                    
                    # If not all documents were inserted, raise an error
                    if successful_inserts != len(data):
                        raise ValueError(f"Failed to import all rows. Expected {len(data)} rows, but only imported {successful_inserts}.")
                    
                    print(f"\nImport Summary:")
                    # print(f"✓ Successfully imported all {successful_inserts} DTC codes")
                    
                except Exception as e:
                    # If there was a bulk write error, provide details
                    if hasattr(e, 'details'):
                        successful_inserts = e.details.get("nInserted", 0)
                        errors = e.details.get("writeErrors", [])
                        error_details = "\n".join([f"Row {err['index']}: {err['errmsg']}" for err in errors[:5]])
                        if len(errors) > 5:
                            error_details += f"\n... and {len(errors) - 5} more errors"
                        
                        raise ValueError(
                            f"Failed to import all rows. Expected {len(data)} rows, but only imported {successful_inserts}.\n"
                            f"First few errors:\n{error_details} {headers}"
                        )
                    raise
            else:
                raise ValueError("No valid data found in Excel file")
                
            return successful_inserts
            
        except Exception as e:
            print(f"Error importing Excel data: {str(e)}")
            raise

    @classmethod
    async def verify_and_update_data(cls):
        """Verify database count matches Excel and update if needed"""
        try:
            excel_count = cls.count_excel_rows()
            db_count = await cls.count_documents()
            
            print(f"\nVerifying DTC data:")
            print(f"Excel rows: {excel_count}")
            print(f"Database documents: {db_count}")
            
            if excel_count != db_count:
                print("\n⚠️  Data mismatch detected. Re-importing DTC data...")
                await cls.import_excel_data(force_update=True)
                new_count = await cls.count_documents()
                print(f"✓ Database updated. New count: {new_count}")
            else:
                print("✓ Data verification passed")
            
        except Exception as e:
            print(f"Error during data verification: {str(e)}")
            raise

    @classmethod
    async def connect(cls):
        """Connect to MongoDB and initialize database"""
        try:
            # Connect to MongoDB
            cls.client = AsyncIOMotorClient(DatabaseConfig.uri)
            cls.db = cls.client[DatabaseConfig.name]
            
            # Drop and recreate collection to apply new schema
            if "dtc_codes" in await cls.db.list_collection_names():
                print("Dropping existing collection to apply new schema...")
                await cls.db.dtc_codes.drop()
            
            print("Creating collection with updated schema...")
            await cls.db.create_collection(
                "dtc_codes",
                validator=create_schema_validation()
            )
            await create_indexes(cls.db.dtc_codes)
            cls.collection = cls.db.dtc_codes
            
            # Import data
            print("Importing DTC codes from Excel...")
            await cls.import_excel_data()
            
            print(f"\nConnected to MongoDB - Database: {DatabaseConfig.name}")
            
        except Exception as e:
            print(f"Error connecting to MongoDB: {e}")
            raise e

    @classmethod
    async def close(cls):
        """Close MongoDB connection"""
        if cls.client:
            cls.client.close()
            print(f"Closed connection to database: {DatabaseConfig.name}")

    @classmethod
    async def is_connected(cls):
        """Check if the database is connected"""
        return cls.client is not None

    @classmethod
    async def get_by_code(cls, dtc_code: str):
        """Get DTC information by code"""
        return await cls.collection.find_one({"DTC": dtc_code})

    @classmethod
    async def get_severity(cls, dtc_code: str):
        """Get severity for a specific DTC code"""
        result = await cls.collection.find_one(
            {"DTC": dtc_code},
            {"SEVERITY": 1, "_id": 0}
        )
        return result["SEVERITY"] if result else None

    @classmethod
    async def count_documents(cls):
        """Get total number of documents"""
        return await cls.collection.count_documents({})

    @classmethod
    async def get_latest_document(cls):
        """Get the latest document"""
        document = await cls.collection.find_one(sort=[("_id", -1)])
        if document and "_id" in document:
            document["_id"] = str(document["_id"])
        return document
    
# Instance to be used in FastAPI app
dtc_db = DTCDatabase()

# Test connection
async def test_connection():
    try:
        await dtc_db.connect()
        count = await dtc_db.count_documents()
        print(f"Total DTC codes: {count}")
        await dtc_db.close()
        print("Connection test completed successfully")
    except Exception as e:
        print(f"Connection test failed: {e}")
        raise e

if __name__ == "__main__":
    import asyncio
    asyncio.run(test_connection())